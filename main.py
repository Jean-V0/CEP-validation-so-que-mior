
import pandas as pd
import requests
import time
import json
import os
import unicodedata
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────
ARQUIVO_ENTRADA = "cep2306.xlsx"
ARQUIVO_SAIDA   = "resultado_validacao_cep.xlsx"
CACHE_FILE      = "cache_cep.json"
COL_CEP         = "CEP"
COL_CIDADE      = "NM_Cidade"
COL_UF          = "ID_UF"
TIMEOUT         = 10
SALVAR_A_CADA   = 200
# ──────────────────────────────────────────────


def normalizar(texto: str) -> str:
    if not texto:
        return ""
    nfkd = unicodedata.normalize("NFKD", str(texto))
    sem_acento = "".join(c for c in nfkd if not unicodedata.combining(c))
    return sem_acento.upper().strip()


def carregar_cache() -> dict:
    if os.path.exists(CACHE_FILE):
        with open(CACHE_FILE, "r", encoding="utf-8") as f:
            cache = json.load(f)
        print(f"Cache carregado: {len(cache):,} CEPs ja consultados.")
        return cache
    return {}


def salvar_cache(cache: dict):
    with open(CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)


def _get(url):
    try:
        return requests.get(url, timeout=TIMEOUT)
    except requests.exceptions.RequestException:
        return None


def consultar_api(cep: str):
    """
    Tenta BrasilAPI -> ViaCEP -> OpenCEP.
    Retorna dict com localidade/uf, None (CEP invalido) ou "TIMEOUT".
    """
    # 1. BrasilAPI
    r = _get(f"https://brasilapi.com.br/api/cep/v1/{cep}")
    if r is not None:
        if r.status_code == 200:
            d = r.json()
            return {"localidade": normalizar(d.get("city", "")),
                    "uf":         normalizar(d.get("state", ""))}
        if r.status_code == 404:
            return None  # CEP invalido confirmado

    # 2. ViaCEP
    r = _get(f"https://viacep.com.br/ws/{cep}/json/")
    if r is not None:
        if r.status_code == 200:
            d = r.json()
            if "erro" not in d:
                return {"localidade": normalizar(d.get("localidade", "")),
                        "uf":         normalizar(d.get("uf", ""))}
            return None  # CEP invalido confirmado
        if r.status_code in (400, 404):
            return None

    # 3. OpenCEP
    r = _get(f"https://opencep.com/v1/{cep}")
    if r is not None:
        if r.status_code == 200:
            d = r.json()
            return {"localidade": normalizar(d.get("localidade", "")),
                    "uf":         normalizar(d.get("uf", ""))}
        if r.status_code == 404:
            return None

    return "TIMEOUT"


def consultar_cep(cep_raw: str, cache: dict):
    cep = cep_raw.strip().replace("-", "").replace(".", "").replace(" ", "").zfill(8)
    if cep in cache:
        return cache[cep]

    resultado = consultar_api(cep)

    # So salva no cache se teve resposta definitiva (nao timeout)
    if resultado != "TIMEOUT":
        cache[cep] = resultado

    time.sleep(0.05)
    return resultado


def main():
    print("=" * 60)
    print("  VALIDADOR DE CEP  --  BrasilAPI / ViaCEP / OpenCEP")
    print("=" * 60)

    if not os.path.exists(ARQUIVO_ENTRADA):
        print(f"\nArquivo '{ARQUIVO_ENTRADA}' nao encontrado.")
        return

    print(f"\nLendo {ARQUIVO_ENTRADA}...")
    df = pd.read_excel(ARQUIVO_ENTRADA, dtype=str)
    df.fillna("", inplace=True)
    total = len(df)

    for col in [COL_CEP, COL_CIDADE, COL_UF]:
        if col not in df.columns:
            print(f"Coluna '{col}' nao encontrada.")
            return

    cache = carregar_cache()
    ceps_unicos = df[COL_CEP].unique()
    pendentes   = [c for c in ceps_unicos
                   if c.strip().replace("-", "").zfill(8) not in cache]

    print(f"{total:,} linhas | {len(ceps_unicos):,} CEPs unicos")
    print(f"Ja no cache : {len(ceps_unicos) - len(pendentes):,}")
    print(f"A consultar : {len(pendentes):,}\n")

    timeouts = 0
    for i, cep in enumerate(pendentes, 1):
        resultado = consultar_cep(cep, cache)
        if resultado == "TIMEOUT":
            timeouts += 1
            print(f"  TIMEOUT: {cep} -- sera retentado na proxima execucao")

        if i % SALVAR_A_CADA == 0:
            salvar_cache(cache)
            pct = i / len(pendentes) * 100
            print(f"  [{i:>6,}/{len(pendentes):,}] {pct:.1f}%  (timeouts: {timeouts})")

    salvar_cache(cache)

    if timeouts:
        print(f"\nATENCAO: {timeouts} CEPs nao responderam em nenhuma API.")
        print("Execute o script novamente para continuar (cache preservado).\n")

    # Validacao
    print(f"Validando {total:,} linhas...")

    status_col, detalhe_col, cidade_api_col, uf_api_col = [], [], [], []

    for _, row in df.iterrows():
        cep_raw = row[COL_CEP]
        cep_key = cep_raw.strip().replace("-", "").replace(".", "").replace(" ", "").zfill(8)

        if cep_key not in cache:
            status_col.append("NAO_CONSULTADO")
            detalhe_col.append("Timeout -- re-execute o script")
            cidade_api_col.append("")
            uf_api_col.append("")
            continue

        dados = cache[cep_key]

        if dados is None:
            status_col.append("CEP_INVALIDO")
            detalhe_col.append(f"CEP '{cep_raw}' nao encontrado nos Correios")
            cidade_api_col.append("")
            uf_api_col.append("")
            continue

        cidade_api = dados.get("localidade", "")
        uf_api     = dados.get("uf", "")
        cidade_api_col.append(cidade_api)
        uf_api_col.append(uf_api)

        erros = []
        cidade_plan = normalizar(row[COL_CIDADE])
        if cidade_plan and cidade_api and cidade_plan != cidade_api:
            erros.append(f"Cidade: planilha='{cidade_plan}' | Correios='{cidade_api}'")
        uf_plan = normalizar(row[COL_UF])
        if uf_plan and uf_api and uf_plan != uf_api:
            erros.append(f"UF: planilha='{uf_plan}' | Correios='{uf_api}'")

        if erros:
            status_col.append("DIVERGENCIA")
            detalhe_col.append(" | ".join(erros))
        else:
            status_col.append("OK")
            detalhe_col.append("")

    df["STATUS_CEP"]      = status_col
    df["CIDADE_CORREIOS"] = cidade_api_col
    df["UF_CORREIOS"]     = uf_api_col
    df["DETALHE_ERRO"]    = detalhe_col

    # Excel formatado
    print("Gerando Excel...")
    df.to_excel(ARQUIVO_SAIDA, index=False)
    wb = load_workbook(ARQUIVO_SAIDA)
    ws = wb.active

    verde    = PatternFill("solid", fgColor="C6EFCE")
    vermelho = PatternFill("solid", fgColor="FFC7CE")
    amarelo  = PatternFill("solid", fgColor="FFEB9C")
    cinza    = PatternFill("solid", fgColor="D9D9D9")

    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="2F5496")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    n = ws.max_column
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        s = row[n - 4].value or ""
        fill = (verde    if s == "OK"           else
                vermelho if s == "CEP_INVALIDO" else
                amarelo  if s == "DIVERGENCIA"  else cinza)
        for cell in row:
            cell.fill = fill

    for ci, w in {n-3: 20, n-2: 25, n-1: 15, n: 65}.items():
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"
    wb.save(ARQUIVO_SAIDA)

    ok          = status_col.count("OK")
    invalidos   = status_col.count("CEP_INVALIDO")
    divergentes = status_col.count("DIVERGENCIA")
    nao_consul  = status_col.count("NAO_CONSULTADO")

    print("\n" + "=" * 60)
    print(f"  RESULTADO -- {total:,} linhas")
    print("=" * 60)
    print(f"  OK              : {ok:>8,}  ({ok/total*100:.1f}%)")
    print(f"  CEP invalido    : {invalidos:>8,}  ({invalidos/total*100:.1f}%)")
    print(f"  Divergencia     : {divergentes:>8,}  ({divergentes/total*100:.1f}%)")
    if nao_consul:
        print(f"  Nao consultado  : {nao_consul:>8,}  (re-execute para completar)")
    print("=" * 60)
    print(f"\nArquivo: {ARQUIVO_SAIDA}")
    print("  Verde = OK | Vermelho = CEP invalido | Amarelo = Divergencia")
    if nao_consul:
        print("  Cinza = Timeout -- re-execute para completar")

    if invalidos + divergentes > 0:
        erradas = df[df["STATUS_CEP"].isin(["CEP_INVALIDO", "DIVERGENCIA"])][
            [COL_CEP, COL_CIDADE, COL_UF, "STATUS_CEP", "DETALHE_ERRO"]
        ].head(30)
        print("\nPrimeiras linhas com problema:")
        print(erradas.to_string())


if __name__ == "__main__":
    main()